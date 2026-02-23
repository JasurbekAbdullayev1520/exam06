# students/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.urls import reverse_lazy
from django.db.models import Q, Count, Avg, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone

from .models import Student, StudentDocument
from .forms import StudentForm, StudentSearchForm, StudentDocumentForm, BulkActionForm

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class StudentListView(LoginRequiredMixin, ListView):
    """
    List view for students with search and filter
    """
    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Student.objects.select_related().prefetch_related('documents')
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(student_id__icontains=search) |
                Q(phone_number__icontains=search)
            )
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by gender
        gender = self.request.GET.get('gender')
        if gender:
            queryset = queryset.filter(gender=gender)
        
        # Filter by city
        city = self.request.GET.get('city')
        if city:
            queryset = queryset.filter(city__icontains=city)
        
        # Filter by enrollment date range
        date_from = self.request.GET.get('enrollment_date_from')
        date_to = self.request.GET.get('enrollment_date_to')
        
        if date_from:
            queryset = queryset.filter(enrollment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(enrollment_date__lte=date_to)
        
        # Ordering
        ordering = self.request.GET.get('ordering', '-created_at')
        if ordering:
            queryset = queryset.order_by(ordering)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = StudentSearchForm(self.request.GET)
        context['total_count'] = Student.objects.count()
        context['active_count'] = Student.objects.filter(status='active').count()
        context['inactive_count'] = Student.objects.filter(status='inactive').count()
        return context


class StudentDetailView(LoginRequiredMixin, DetailView):
    """
    Detail view for single student
    """
    model = Student
    template_name = 'students/student_detail.html'
    context_object_name = 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object
        
        # Get enrollments
        context['enrollments'] = student.enrollments.select_related(
            'course'
        ).order_by('-enrollment_date')
        
        # Get payments
        try:
            from payments.models import Payment
            context['payments'] = Payment.objects.filter(
                student=student
            ).order_by('-created_at')[:10]
            
            context['total_paid'] = student.get_total_paid()
            context['total_debt'] = student.get_total_debt()
        except:
            pass
        
        # Get attendance
        try:
            context['attendance_percentage'] = student.get_attendance_percentage()
        except:
            context['attendance_percentage'] = 0
        
        # Get documents
        context['documents'] = student.documents.order_by('-uploaded_at')
        
        # Get active courses count
        context['active_courses_count'] = student.get_active_courses().count()
        
        return context


class StudentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Create view for new student
    """
    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('students:student_list')
    permission_required = 'students.add_student'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Talaba muvaffaqiyatli qo\'shildi!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Iltimos, xatolarni to\'g\'rilang.')
        return super().form_invalid(form)


class StudentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Update view for existing student
    """
    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('students:student_list')
    permission_required = 'students.change_student'
    
    def form_valid(self, form):
        messages.success(self.request, 'Talaba ma\'lumotlari yangilandi!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Iltimos, xatolarni to\'g\'rilang.')
        return super().form_invalid(form)


class StudentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Delete view for student
    """
    model = Student
    template_name = 'students/student_confirm_delete.html'
    success_url = reverse_lazy('students:student_list')
    permission_required = 'students.delete_student'
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Talaba muvaffaqiyatli o\'chirildi!')
        return super().delete(request, *args, **kwargs)


# Document Views
@login_required
def student_document_upload(request, pk):
    """Upload document for student"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.student = student
            document.uploaded_by = request.user
            document.save()
            
            messages.success(request, 'Hujjat muvaffaqiyatli yuklandi!')
            return redirect('students:student_detail', pk=pk)
    else:
        form = StudentDocumentForm()
    
    return render(request, 'students/document_upload.html', {
        'form': form,
        'student': student
    })


@login_required
def student_document_delete(request, pk, doc_id):
    """Delete student document"""
    student = get_object_or_404(Student, pk=pk)
    document = get_object_or_404(StudentDocument, pk=doc_id, student=student)
    
    if request.method == 'POST':
        document.file.delete()
        document.delete()
        messages.success(request, 'Hujjat o\'chirildi!')
    
    return redirect('students:student_detail', pk=pk)


# Export Views
@login_required
@permission_required('students.view_student')
def export_students_excel(request):
    """Export students to Excel"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Talabalar'
    
    # Header style
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    # Headers
    headers = [
        'ID', 'Ism', 'Familiya', 'Email', 'Telefon',
        'Tug\'ilgan sana', 'Yosh', 'Jins', 'Shahar',
        'Holat', 'Ro\'yxatdan o\'tgan sana'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get filtered students
    queryset = Student.objects.all()
    
    # Apply filters from request
    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)
    
    # Write data
    for row_num, student in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=student.student_id)
        ws.cell(row=row_num, column=2, value=student.first_name)
        ws.cell(row=row_num, column=3, value=student.last_name)
        ws.cell(row=row_num, column=4, value=student.email)
        ws.cell(row=row_num, column=5, value=student.phone_number)
        ws.cell(row=row_num, column=6, value=student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '')
        ws.cell(row=row_num, column=7, value=student.get_age())
        ws.cell(row=row_num, column=8, value=student.get_gender_display())
        ws.cell(row=row_num, column=9, value=student.city)
        ws.cell(row=row_num, column=10, value=student.get_status_display())
        ws.cell(row=row_num, column=11, value=student.enrollment_date.strftime('%Y-%m-%d'))
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=talabalar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# Bulk Actions
@login_required
@permission_required('students.change_student')
def bulk_action(request):
    """Perform bulk actions on students"""
    
    if request.method == 'POST':
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_students')
        
        if not selected_ids:
            messages.warning(request, 'Talabalarni tanlang!')
            return redirect('students:student_list')
        
        students = Student.objects.filter(id__in=selected_ids)
        
        if action == 'activate':
            students.update(status='active')
            messages.success(request, f'{students.count()} ta talaba faollashtirildi!')
        
        elif action == 'deactivate':
            students.update(status='inactive')
            messages.success(request, f'{students.count()} ta talaba faolsizlantirildi!')
        
        elif action == 'delete':
            count = students.count()
            students.delete()
            messages.success(request, f'{count} ta talaba o\'chirildi!')
        
        elif action == 'export':
            return export_students_excel(request)
    
    return redirect('students:student_list')


# Statistics View
@login_required
def student_statistics(request):
    """View student statistics"""
    
    total_students = Student.objects.count()
    active_students = Student.objects.filter(status='active').count()
    
    # Gender distribution
    male_count = Student.objects.filter(gender='M').count()
    female_count = Student.objects.filter(gender='F').count()
    
    # City distribution
    city_stats = Student.objects.values('city').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Monthly registration data
    from django.db.models.functions import TruncMonth
    monthly_registrations = Student.objects.annotate(
        month=TruncMonth('enrollment_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    context = {
        'total_students': total_students,
        'active_students': active_students,
        'male_count': male_count,
        'female_count': female_count,
        'city_stats': city_stats,
        'monthly_registrations': monthly_registrations,
    }
    
    return render(request, 'students/statistics.html', context)


# AJAX Views
@login_required
def check_email_availability(request):
    """Check if email is available"""
    email = request.GET.get('email')
    student_id = request.GET.get('student_id')
    
    if email:
        exists = Student.objects.filter(email=email).exclude(
            student_id=student_id
        ).exists()
        
        return JsonResponse({'available': not exists})
    
    return JsonResponse({'available': False})


@login_required
def student_quick_search(request):
    """Quick search for students (AJAX)"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    students = Student.objects.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(email__icontains=query) |
        Q(student_id__icontains=query)
    )[:10]
    
    results = [{
        'id': s.id,
        'student_id': s.student_id,
        'name': s.get_full_name(),
        'email': s.email,
        'status': s.get_status_display(),
    } for s in students]
    
    return JsonResponse({'results': results})